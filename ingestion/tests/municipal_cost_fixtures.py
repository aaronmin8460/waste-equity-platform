"""SYNTHETIC, TEST-ONLY workbook builders for the municipal waste-cost parser.

Every workbook produced here is fabricated in ``tmp_path``. None of it is
official data and none of it is ever presented as such. The real 2024 municipal
disclosure workbooks are Git-ignored local data and are never read, copied, or
committed by the test suite.

The header spines mirror the seven layouts the Step 1 audit found, so a parser
change that breaks a real layout breaks a test here first.
"""

from __future__ import annotations

import re
import shutil
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook

# --- DATA_A header spines (the trailing 최종 처리시설/처리방식/비고 block is what
# --- the parser locates by name; the quantity pair columns vary in count) ------

HEADER_TAIL = ["최종 처리시설", "처리방식", "비고"]

# Family 2 — 9 columns, one [label, value] quantity pair.
HEADERS_FAMILY_2 = [
    "년도",
    "기관명",
    "계약명",
    "총 금액(총 지급액)",
    "월별 생활 쓰레기 반출량",
    None,
    *HEADER_TAIL,
]
# Family 2 at 11 columns (two pairs) and 15 columns (four pairs).
HEADERS_FAMILY_2_WIDE = [
    "년도",
    "기관명",
    "계약명",
    "총 금액(총 지급액)",
    "월별 생활 쓰레기 반출량",
    None,
    None,
    None,
    *HEADER_TAIL,
]
HEADERS_FAMILY_2_WIDEST = [
    "년도",
    "기관명",
    "계약명",
    "총 금액(총 지급액)",
    "월별 생활 쓰레기 반출량",
    None,
    None,
    None,
    None,
    None,
    None,
    None,
    *HEADER_TAIL,
]
# Family 3 — family 2 plus a per-month payment ledger.
HEADERS_FAMILY_3 = [
    "년도",
    "기관명",
    "계약명",
    "총 금액(총 지급액)",
    "월별 생활 쓰레기 반출량",
    "(수치)",
    "(반출경로2)",
    "(수치)",
    *HEADER_TAIL,
    "지급월",
    "구분",
    "월별 지급액 합계(원)",
]
# Family 4 — numbered pair headers.
HEADERS_FAMILY_4 = [
    "년도",
    "기관명",
    "계약명",
    "총 금액(총 지급액)",
    "월별 생활 쓰레기 반출량",
    "(E의 수치)",
    "(두 번째 반출 경로 항목명)",
    "(G의 수치)",
    *HEADER_TAIL,
]
# Family 5 — three destination pairs and the 총 금액(원) payment header.
HEADERS_FAMILY_5 = [
    "년도",
    "기관명",
    "계약명",
    "총 금액(원)",
    "월별 반출량 항목(매립)",
    "반출량(톤)",
    "월별 반출량 항목(소각)",
    "반출량(톤)",
    "월별 반출량 항목(청라크린넷)",
    "반출량(톤)",
    *HEADER_TAIL,
]
# Family 6 — family 2 spine whose every quantity cell is a source dash.
HEADERS_FAMILY_6 = [
    "년도",
    "기관명",
    "계약명",
    "총 금액(총 지급액)",
    "월별 생활 쓰레기 반출량",
    "(수치)",
    "(두 번째 반출 경로)",
    "(수치)",
    *HEADER_TAIL,
]
# Family 7 — tonne-labelled pair headers.
HEADERS_FAMILY_7 = [
    "년도",
    "기관명",
    "계약명",
    "총 금액(총 지급액)",
    "월별 생활 쓰레기 반출량(항목명)",
    "반출량(톤)",
    "반출경로2(항목명)",
    "반출량2(톤)",
    *HEADER_TAIL,
]

# Family 1 — the fixed DATA_B tonnage grid.
HEADERS_DATA_B = ["구분", *[f"{month}월" for month in range(1, 13)], "계"]


def write_workbook(
    path: Path,
    headers: list[Any],
    rows: list[list[Any]],
    *,
    sheet_name: str = "Sheet1",
    merges: list[str] | None = None,
    extra_empty_sheets: tuple[str, ...] = (),
) -> Path:
    """Write one synthetic workbook. Returns the path for convenience."""

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    for merge in merges or ():
        sheet.merge_cells(merge)
    for name in extra_empty_sheets:
        workbook.create_sheet(name)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def contract_row(
    *,
    year: Any = 2024,
    organisation: str = "테스트시청",
    name: str = "테스트시 생활폐기물 수집·운반 대행용역(1구역)",
    payment: Any = 1_000_000,
    pairs: list[Any] | None = None,
    facility: str = "테스트소각장;테스트매립지",
    method: str = "소각·매립",
    note: str = "",
    trailing: list[Any] | None = None,
) -> list[Any]:
    """One DATA_A contract header row for the given number of quantity pairs."""

    return [
        year,
        organisation,
        name,
        payment,
        *(pairs or [None, None]),
        facility,
        method,
        note,
        *(trailing or []),
    ]


def quantity_row(
    pairs: list[Any], *, pair_offset: int = 4, trailing: list[Any] | None = None
) -> list[Any]:
    """A continuation row carrying only quantity pair cells (and optional tail)."""

    return [None] * pair_offset + list(pairs) + [None, None, None] + list(trailing or [])


def monthly_pairs(label: str, values: list[Any]) -> list[list[Any]]:
    """Twelve ``[label: N월, value]`` continuation-row pairs."""

    return [[f"{label}: {month}월", value] for month, value in enumerate(values, start=1)]


def inject_cached_formula(
    path: Path,
    cell: str,
    formula: str,
    cached: Any,
    *,
    sheet_file: str = "xl/worksheets/sheet1.xml",
) -> None:
    """Rewrite one cell as a formula **with** a cached value.

    openpyxl cannot author a cached formula result, but every real source
    workbook has one (the audit found 0 formulas without a cached value), so the
    fixture is patched at the XML level. The parser reads with ``data_only=True``
    and must see the cached number.
    """

    temporary = path.with_suffix(".patched.xlsx")
    pattern = re.compile(rf'<c r="{cell}"[^>]*>.*?</c>', re.DOTALL)
    replacement = f'<c r="{cell}"><f>{formula}</f><v>{cached}</v></c>'
    with (
        zipfile.ZipFile(path) as source,
        zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as target,
    ):
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == sheet_file:
                xml = data.decode("utf-8")
                xml, count = pattern.subn(replacement, xml, count=1)
                if count != 1:  # pragma: no cover - fixture authoring error
                    raise AssertionError(f"cell {cell} not found in {sheet_file}")
                data = xml.encode("utf-8")
            target.writestr(item, data)
    shutil.move(str(temporary), str(path))


def data_b_rows(
    general: list[Any],
    food: list[Any],
    recycling: list[Any],
    totals: tuple[Any, Any, Any, Any] | None = None,
    combined: list[Any] | None = None,
) -> list[list[Any]]:
    """The four ``일반 / 음식물 / 재활용 / 계`` rows of a DATA_B grid."""

    general_total, food_total, recycling_total, combined_total = totals or (
        None,
        None,
        None,
        None,
    )
    if combined is None:
        combined = [_add(a, b, c) for a, b, c in zip(general, food, recycling, strict=True)]
    return [
        ["일반", *general, general_total],
        ["음식물", *food, food_total],
        ["재활용", *recycling, recycling_total],
        ["계", *combined, combined_total],
    ]


def _add(*values: Any) -> Any:
    numbers = [value for value in values if isinstance(value, int | float)]
    return sum(numbers) if numbers else "-"
