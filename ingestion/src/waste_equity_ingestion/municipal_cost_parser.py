"""Deterministic openpyxl parser for the 2024 municipal waste-cost workbooks.

Pure and side-effect free: it opens each workbook **read-only**, never resaves or
rewrites a source file, uses no OCR and no LibreOffice, performs no network or
database access, and returns the same result for the same bytes every time.

Two source families, seven audited worksheet layouts:

``DATA_A`` — contract rows carrying the payment in the 총 금액(총 지급액) /
총 금액(원) / ``<year>년 금액`` column, with a variable number of
``[label, value]`` quantity pair columns between the payment column and the
trailing ``최종 처리시설 / 처리방식 / 비고`` block. Every column is located **by
header name**, never by index, so the 7/9/11/13/14/15-column variants share one
code path, and the year header is accepted as either 년도 or 연도. Locating the
payment column never decides what the amount means: the actual-paid / contract-award
/ budget-estimate distinction stays with :func:`classify_payment_type`, which only
ever reclassifies on wording the source itself wrote.

An amount can be a genuine actual payment and still not belong in this
indicator's numerator. :func:`classify_contract_accounting_basis` answers the
separate question of *what kind of money* a contract is — a 수도권매립지
반입수수료 and a 위탁처리 용역 are both really paid, and neither is a
collection-and-transport payment. The two axes stay orthogonal, so an excluded
contract keeps its true ``payment_type``, its amount, and its source row; only
``is_primary_numerator_eligible`` turns False, under
``NON_COLLECTION_TRANSPORT_BASIS``.

An amount delivered as formatted text ("6,556,861,120원") is read as that exact
number by :func:`parse_money_text` under an anchored pattern; anything that is not
a well-formed currency literal stays missing rather than becoming a number.

``DATA_B`` — a fixed ``구분 / 1월…12월 / 계`` × ``일반, 음식물, 재활용, 계``
tonnage grid. It carries **no monetary field anywhere**, and this parser
structurally cannot produce a payment from it: :func:`parse_workbook` returns an
empty contract list for every DATA_B file.

Missing is never zero. A blank cell, a source ``-`` placeholder, and a source
"자료 없음"/"미제공"/"확인 불가" text each keep ``quantity_value = None`` under
their own ``value_state``; only a real numeric 0 becomes ``MEASURED_ZERO``. A
``계`` cell holding 0 produced by ``=SUM()`` over twelve dashes is reported as
``SOURCE_DASH_NO_DATA`` + ``ZERO_TOTAL_QUANTITY``, never as a measured zero.

Every string that is compared, matched, or keyed is normalized to Unicode NFC
first — macOS hands back Korean filenames and cell text in NFD and a naive
comparison silently matches nothing.
"""

from __future__ import annotations

import datetime
import hashlib
import re
from collections.abc import Iterator, Sequence
from dataclasses import dataclass, field
from decimal import ROUND_HALF_EVEN, Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from waste_equity_backend.analysis.municipal_cost import nfc
from waste_equity_backend.models.municipal_cost import (
    ATTRIBUTION_MUNICIPAL_TOTAL_REPEATED,
    ATTRIBUTION_MUNICIPAL_TOTAL_SINGLE,
    ATTRIBUTION_PER_CONTRACT,
    CLASS_DATA_A_PARTIAL_WASTE_SCOPE,
    CLASS_DATA_A_PAYMENT_AND_QUANTITY,
    CLASS_DATA_A_PAYMENT_ONLY,
    CLASS_DATA_A_QUANTITY_ONLY,
    CLASS_DATA_B_QUANTITY_VALIDATION,
    CLASS_EMPTY_OR_NO_DATA,
    CLASS_UNSUPPORTED_LAYOUT,
    COMPLETENESS_COMPLETE,
    COMPLETENESS_PARTIAL,
    COMPLETENESS_UNUSABLE,
    CONTRACT_BASIS_COLLECTION_TRANSPORT,
    CONTRACT_BASIS_COLLECTION_TRANSPORT_WITH_TREATMENT,
    CONTRACT_BASIS_FACILITY_INBOUND_FEE,
    CONTRACT_BASIS_TREATMENT_SERVICE,
    DATASET_ROLE_A,
    DATASET_ROLE_B,
    GEOGRAPHIC_SCOPE_SUB,
    GEOGRAPHIC_SCOPE_WHOLE,
    LAYOUT_DATA_A_CONTRACT_DASH_ONLY_QUANTITY,
    LAYOUT_DATA_A_CONTRACT_DESTINATION_SPLIT,
    LAYOUT_DATA_A_CONTRACT_NUMBERED_PAIRS,
    LAYOUT_DATA_A_CONTRACT_QUANTITY_PAIRS,
    LAYOUT_DATA_A_CONTRACT_TONNE_LABELLED_PAIRS,
    LAYOUT_DATA_A_CONTRACT_WITH_PAYMENT_LEDGER,
    LAYOUT_DATA_B_MONTHLY_CATEGORY_GRID,
    LAYOUT_UNSUPPORTED,
    NUMERATOR_ELIGIBLE_BASES,
    PAYMENT_ACTUAL_PAID,
    PAYMENT_BUDGET_ESTIMATE,
    PAYMENT_CONTRACT_AWARD,
    QUANTITY_DECIMAL_PLACES,
    QUANTITY_PERIOD_ANNUAL,
    QUANTITY_PERIOD_MONTHLY,
    QUANTITY_PERIOD_MONTHLY_AVERAGE,
    QUANTITY_PERIOD_UNSPECIFIED,
    REASON_AMBIGUOUS_SOURCE_MEANING,
    REASON_INCONSISTENT_TOTAL,
    REASON_MISSING_QUANTITY,
    REASON_MIXED_REFERENCE_YEARS,
    REASON_NON_COLLECTION_TRANSPORT_BASIS,
    REASON_PARTIAL_GEOGRAPHIC_SCOPE,
    REASON_PARTIAL_PERIOD_COVERAGE,
    REASON_PARTIAL_WASTE_SCOPE,
    REASON_PAYMENT_PERIOD_COVERAGE_INCOMPLETE,
    REASON_UNSUPPORTED_SOURCE_LAYOUT,
    REASON_ZERO_TOTAL_QUANTITY,
    VALUE_BLANK,
    VALUE_MEASURED,
    VALUE_MEASURED_ZERO,
    VALUE_SOURCE_DASH_NO_DATA,
    VALUE_SOURCE_TEXT_NO_DATA,
    WASTE_CATEGORY_BULKY,
    WASTE_CATEGORY_COMBINED,
    WASTE_CATEGORY_FOOD,
    WASTE_CATEGORY_GENERAL,
    WASTE_CATEGORY_RECYCLING,
    WASTE_CATEGORY_UNKNOWN,
    WASTE_SCOPE_ALL_MUNICIPAL,
    WASTE_SCOPE_FOOD_ONLY,
    WASTE_SCOPE_GENERAL_ONLY,
    WASTE_SCOPE_MIXED_PARTIAL,
    WASTE_SCOPE_RECYCLING_ONLY,
    WASTE_SCOPE_UNKNOWN,
)


class MunicipalCostParseError(RuntimeError):
    """Raised only for an unreadable workbook; layout problems are reported."""


# --- header vocabulary (all matched after NFC + whitespace normalization) ---
# 년도 and 연도 are two orthographic spellings of the same word and are used
# interchangeably by different suppliers (오산시 writes 연도). Both are matched
# exactly against this enumerated tuple, never by substring.
HEADER_YEAR_VARIANTS = ("년도", "연도")
HEADER_ORGANISATION = "기관명"
HEADER_CONTRACT_NAME = "계약명"
HEADER_PAYMENT_PREFIX = "총 금액"
HEADER_FACILITY = "최종 처리시설"
HEADER_METHOD = "처리방식"
HEADER_NOTE = "비고"
HEADER_LEDGER_MONTH = "지급월"
HEADER_LEDGER_KIND = "구분"
HEADER_LEDGER_AMOUNT_PREFIX = "월별 지급액 합계"

DATA_B_HEADER_CATEGORY = "구분"
DATA_B_HEADER_TOTAL = "계"
DATA_B_CATEGORY_ROWS = ("일반", "음식물", "재활용", "계")
DATA_B_CATEGORY_TO_WASTE = {
    "일반": WASTE_CATEGORY_GENERAL,
    "음식물": WASTE_CATEGORY_FOOD,
    "재활용": WASTE_CATEGORY_RECYCLING,
    "계": WASTE_CATEGORY_COMBINED,
}

TOTAL_ROW_TOKENS = ("합계", "총계", "소계", "계")


def payment_header_variants(reference_year: int) -> tuple[str, ...]:
    """Exact payment-column header spellings, beyond the 총 금액… prefix.

    ``오산시`` labels the same column ``2024년 금액``. The form is enumerated and
    anchored to the reference year rather than matched as "a header containing
    금액", so ``계약금액`` / ``낙찰금액`` / ``예산액`` can never be picked up.

    Locating a column decides **where** the amount is, never **what it means**:
    payment semantics remain with :func:`classify_payment_type`, which reclassifies
    away from ``ACTUAL_PAID_AMOUNT`` only when the source itself writes 계약금액 or
    예산·예상금액. This function therefore adds no new way for a non-payment figure
    to enter the numerator.
    """

    return (f"{reference_year}년 금액",)


# A monetary cell delivered as *text* rather than as a number
# ("6,556,861,120원" — 미추홀구 formats every payment this way). The pattern is
# anchored and strict: well-formed thousands groups or plain digits, an optional
# decimal part, an optional 원 suffix, and nothing else. A dash, a "확인 불가"
# placeholder, a range, or any prose fails to match and the amount stays missing.
# This never produces a zero and never guesses.
_MONEY_TEXT_PATTERN = re.compile(r"^(?P<int>\d{1,3}(?:,\d{3})+|\d+)(?:\.(?P<frac>\d+))?\s*원?$")


def parse_money_text(raw: Any) -> Decimal | None:
    """Exact Decimal for a source-formatted KRW literal held as text; else None."""

    if not isinstance(raw, str):
        return None
    match = _MONEY_TEXT_PATTERN.match(_text(raw))
    if match is None:
        return None
    digits = match.group("int").replace(",", "")
    fraction = match.group("frac")
    try:
        return Decimal(f"{digits}.{fraction}") if fraction else Decimal(digits)
    except InvalidOperation:  # pragma: no cover - the pattern guarantees a number
        return None


# Source placeholders. A dash means "the source declared no data for this cell";
# these prose tokens mean "the source states the data does not exist". Neither is
# ever coerced to zero.
DASH_TOKENS = ("-", "－", "–", "—", "ㅡ", "‐", "‑")
NO_DATA_TEXT_TOKENS = (
    "미제공",
    "자료없음",
    "자료 없음",
    "확인불가",
    "확인 불가",
    "자료부존재",
    "자료 부존재",
    "부존재",
    "해당없음",
    "해당 없음",
    "없음",
)

# Note markers that reclassify the payment away from an actual paid amount. Each
# is a phrase the source itself writes; nothing is inferred from an amount alone.
BUDGET_ESTIMATE_MARKERS = ("예산/예상금액", "실적 정산치 아닌", "예산·예상금액")
CONTRACT_AWARD_MARKERS = (
    "d열은 계약금액",
    "총금액은 계약금액",
    "총 금액은 계약금액",
    "지급액이 아닌 계약금액",
)

# --- accounting-basis vocabulary (see classify_contract_accounting_basis) ---
#
# Every token below is a word the source writes in ``계약명``; nothing is inferred
# from an amount, a destination, or a municipality. All matching happens on
# NFC-normalized, whitespace-collapsed text, which is why the spaced and unspaced
# spellings of the same compound are both enumerated rather than pattern-matched:
# "민간위탁 소각처리 용역" and "민간소각 처리 용역" are the same rule, written two
# ways by two suppliers.

# The indicator's own vocabulary. Any one of these is enough to establish that
# the municipality bought haulage.
COLLECTION_TRANSPORT_TOKENS = ("수집", "운반", "수송", "청소")

# A facility gate fee — a different accounting basis under any wording.
FACILITY_INBOUND_FEE_TOKENS = ("반입수수료", "반입 수수료")

# Treatment bought as a service. Each token pairs an action with 처리, so a bare
# 처리 (which 27 ordinary collection contracts mention) never matches, and neither
# does a destination such as 자원회수시설 or 소각장.
TREATMENT_ONLY_TOKENS = (
    "위탁처리",
    "위탁 처리",
    "처리위탁",
    "처리 위탁",
    "소각처리",
    "소각 처리",
    "선별처리",
    "선별 처리",
    "처리용역",
    "처리 용역",
    "처리대행",
    "처리 대행",
    "대행처리",
    "대행 처리",
)

# Filename annotation markers. The delivered filenames carry the supplier's own
# scope notes ("부평구 - 생활(일반)수집운반만 있고 음식물류·재활용 데이터 없음.xlsx"),
# which are part of the delivered dataset and are the only place some
# municipalities state their waste scope.
FILENAME_GENERAL_ONLY_MARKERS = ("(일반)",)
FILENAME_ONLY_MARKERS = ("만 있고", "만 포함", "만있고")

# A quantity label naming a sub-municipal 읍/면 inside parentheses (옹진군's two
# contracts are 영흥면 / 북도·영흥면 only).
SUB_MUNICIPAL_PATTERN = re.compile(r"\([^)]*[읍면][^)]*\)")

_MONTH_ISO_PATTERN = re.compile(r"^(\d{4})[-.](\d{1,2})$")
_MONTH_KO_PATTERN = re.compile(r"^(\d{1,2})월")
_PART_YEAR_PATTERN = re.compile(r"^\s*(\d{4})\s*[.\-/]")
_DATE_TOKEN_PATTERN = re.compile(r"(\d{4})?\s*[.\-/]?\s*(\d{1,2})\s*[.\-/]\s*(\d{1,2})\s*\.?\s*$")
_COMPANY_SUFFIX_PATTERN = re.compile(r"(\(주\)|\(유\)|㈜|㈜|주식회사|유한회사)")
_CONTRACTOR_LABELS = ("계약상대자", "계약상대방", "업체")


def _text(value: Any) -> str:
    """NFC-normalized, whitespace-collapsed text for a cell value."""

    if value is None:
        return ""
    if isinstance(value, str):
        return re.sub(r"\s+", " ", nfc(value)).strip()
    return nfc(str(value)).strip()


def _is_dash(text: str) -> bool:
    return text in DASH_TOKENS


def _is_no_data_text(text: str) -> bool:
    stripped = text.replace(" ", "")
    return any(token.replace(" ", "") == stripped for token in NO_DATA_TEXT_TOKENS)


# Tonnage is reported to at most four decimals. A cached formula cell can carry
# 11-13 decimals of IEEE-754 noise around a two-decimal source value (남동구's
# 432.4299999999999 *is* 432.43); quantizing to the source's declared precision
# removes the artefact without inventing a tolerance or altering a real value.
QUANTITY_QUANTUM = Decimal(1).scaleb(-QUANTITY_DECIMAL_PLACES)


def quantize_quantity(value: Decimal) -> Decimal:
    return value.quantize(QUANTITY_QUANTUM, rounding=ROUND_HALF_EVEN)


def _decimal(value: Any) -> Decimal | None:
    """Exact Decimal for a numeric cell; None for anything else.

    ``bool`` is excluded explicitly because it is an ``int`` subclass and a
    ``TRUE`` cell must never be read as the number 1.
    """

    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        # str(float) round-trips the shortest exact repr, which is the value the
        # spreadsheet displays; Decimal(float) would import binary noise.
        return Decimal(str(value))
    return None


@dataclass(frozen=True)
class ParsedQuantity:
    """One logical quantity observation read from a worksheet."""

    source_row: int
    source_column: int
    source_label: str
    source_repetition_rows: tuple[int, ...]
    contract_index: int | None
    reference_month: str | None
    quantity_period: str
    waste_category: str
    destination_name: str | None
    treatment_method: str | None
    quantity_value: Decimal | None
    value_state: str
    attribution: str
    limitation_reasons: tuple[str, ...] = ()


@dataclass(frozen=True)
class ParsedContract:
    """One contract header row — the payment grain."""

    source_row: int
    source_row_end: int
    contract_name: str
    contractor_name: str | None
    source_period_label: str | None
    payment_amount_krw: Decimal | None
    payment_source_text: str | None
    payment_type: str
    # What kind of money this is, independent of whether it was actually paid.
    accounting_basis: str
    is_primary_numerator_eligible: bool
    payment_period_start: datetime.date | None
    payment_period_end: datetime.date | None
    waste_scope: str
    geographic_scope: str
    destination_names: tuple[str, ...]
    treatment_methods: tuple[str, ...]
    completeness_status: str
    limitation_reasons: tuple[str, ...]
    source_note: str | None


@dataclass
class ParsedWorkbook:
    """Everything :func:`parse_workbook` could determine about one file."""

    dataset_role: str
    sheet_name: str
    used_range: str
    layout_family: str
    primary_classification: str
    reference_year: int
    source_municipality_names: tuple[str, ...] = ()
    contracts: list[ParsedContract] = field(default_factory=list)
    quantities: list[ParsedQuantity] = field(default_factory=list)
    source_total_krw: Decimal | None = None
    source_total_text: str | None = None
    recomputed_total_krw: Decimal | None = None
    source_notes: list[str] = field(default_factory=list)
    # File-level limitation reasons that concern the *payment* side and therefore
    # may reach the indicator. Quantity-only limitations stay on the quantity rows.
    payment_limitation_reasons: list[str] = field(default_factory=list)
    quantity_limitation_reasons: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def has_numeric_quantity(self) -> bool:
        return any(q.quantity_value is not None for q in self.quantities)

    @property
    def eligible_payment_total(self) -> Decimal | None:
        eligible = [
            c.payment_amount_krw
            for c in self.contracts
            if c.is_primary_numerator_eligible and c.payment_amount_krw is not None
        ]
        return sum(eligible, Decimal(0)) if eligible else None

    @property
    def eligible_contract_count(self) -> int:
        return sum(1 for c in self.contracts if c.is_primary_numerator_eligible)


def sha256_file(path: Path, *, chunk_size: int = 1 << 20) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(chunk_size), b""):
            digest.update(chunk)
    return digest.hexdigest()


# ---------------------------------------------------------------------------
# Shared cell-level classification
# ---------------------------------------------------------------------------


def classify_value(raw: Any, *, label: str = "") -> tuple[Decimal | None, str]:
    """Map one source cell to ``(value, value_state)``.

    The label participates because several workbooks put the placeholder in the
    *label* column and leave the value column empty ("미제공" with no number);
    reading that as ``BLANK`` would lose the source's explicit statement that the
    data does not exist.
    """

    number = _decimal(raw)
    if number is not None:
        number = quantize_quantity(number)
        return (number, VALUE_MEASURED_ZERO if number == 0 else VALUE_MEASURED)

    text = _text(raw)
    if _is_dash(text):
        return (None, VALUE_SOURCE_DASH_NO_DATA)
    if text and _is_no_data_text(text):
        return (None, VALUE_SOURCE_TEXT_NO_DATA)
    if text:
        # Non-numeric prose that is not a recognised placeholder: still not a
        # number, and still never zero.
        return (None, VALUE_SOURCE_TEXT_NO_DATA)

    label_text = _text(label)
    if _is_dash(label_text):
        return (None, VALUE_SOURCE_DASH_NO_DATA)
    if label_text and _is_no_data_text(label_text):
        return (None, VALUE_SOURCE_TEXT_NO_DATA)
    return (None, VALUE_BLANK)


def parse_label_period(label: str, reference_year: int) -> tuple[str | None, str]:
    """Split a quantity label's trailing period token into ``(month, period)``.

    Handles ``… : 1월`` (reference year), ``… : 2023-06`` (explicit year, which is
    what raises MIXED_REFERENCE_YEARS), ``… : 계`` (annual total), ``월평균: …``
    (a monthly *average*, never a total), and anything else as UNSPECIFIED.
    """

    text = _text(label)
    if not text:
        return (None, QUANTITY_PERIOD_UNSPECIFIED)
    if text.startswith("월평균"):
        return (None, QUANTITY_PERIOD_MONTHLY_AVERAGE)
    tail = text.rsplit(":", 1)[-1].strip() if ":" in text else text

    iso = _MONTH_ISO_PATTERN.match(tail)
    if iso:
        year, month = int(iso.group(1)), int(iso.group(2))
        if 1 <= month <= 12:
            return (f"{year:04d}-{month:02d}", QUANTITY_PERIOD_MONTHLY)
    korean = _MONTH_KO_PATTERN.match(tail)
    if korean:
        month = int(korean.group(1))
        if 1 <= month <= 12:
            return (f"{reference_year:04d}-{month:02d}", QUANTITY_PERIOD_MONTHLY)
    if tail in TOTAL_ROW_TOKENS:
        return (None, QUANTITY_PERIOD_ANNUAL)
    return (None, QUANTITY_PERIOD_UNSPECIFIED)


def classify_waste_category(label: str, contract_scope: str) -> str:
    """Waste stream of a quantity row: label wording first, contract scope second."""

    text = _text(label)
    named = [
        ("음식물", WASTE_CATEGORY_FOOD),
        ("재활용", WASTE_CATEGORY_RECYCLING),
        ("대형", WASTE_CATEGORY_BULKY),
    ]
    hits = {category for token, category in named if token in text}
    if "일반" in text or "가연성" in text:
        hits.add(WASTE_CATEGORY_GENERAL)
    if len(hits) == 1:
        return hits.pop()
    if len(hits) > 1:
        return WASTE_CATEGORY_COMBINED
    return {
        WASTE_SCOPE_ALL_MUNICIPAL: WASTE_CATEGORY_COMBINED,
        WASTE_SCOPE_GENERAL_ONLY: WASTE_CATEGORY_GENERAL,
        WASTE_SCOPE_FOOD_ONLY: WASTE_CATEGORY_FOOD,
        WASTE_SCOPE_RECYCLING_ONLY: WASTE_CATEGORY_RECYCLING,
        WASTE_SCOPE_MIXED_PARTIAL: WASTE_CATEGORY_COMBINED,
        WASTE_SCOPE_UNKNOWN: WASTE_CATEGORY_UNKNOWN,
    }.get(contract_scope, WASTE_CATEGORY_UNKNOWN)


def classify_contract_waste_scope(contract_name: str, *, filename_forces_general: bool) -> str:
    """Waste scope of one contract, from the source's own wording.

    ``생활폐기물`` / ``생활쓰레기`` / ``청소`` name the whole municipal stream in
    Korean administrative usage; only an explicit narrowing token —
    ``(일반)``, ``(재활용…)``, ``(음식물…)``, ``가연성``, ``대형폐기물`` — makes a
    contract partial. A name that enumerates two or more streams (서구's
    ``생활·재활용가능자원·음식물류폐기물``) is whole-stream again.
    """

    if filename_forces_general:
        return WASTE_SCOPE_GENERAL_ONLY
    name = _text(contract_name)
    if not name:
        return WASTE_SCOPE_UNKNOWN

    narrow: set[str] = set()
    if "(일반)" in name or "일반쓰레기" in name or "가연성" in name:
        narrow.add(WASTE_SCOPE_GENERAL_ONLY)
    if "음식물" in name:
        narrow.add(WASTE_SCOPE_FOOD_ONLY)
    if "재활용" in name or "미선별플라스틱" in name:
        narrow.add(WASTE_SCOPE_RECYCLING_ONLY)
    if "대형폐기물" in name:
        narrow.add(WASTE_SCOPE_MIXED_PARTIAL)

    if len(narrow) >= 2:
        # Two or more named streams in one contract: the source is describing a
        # combined scope, not a narrower one.
        return WASTE_SCOPE_ALL_MUNICIPAL
    if len(narrow) == 1:
        return narrow.pop()
    if "생활폐기물" in name or "생활쓰레기" in name or "청소" in name or "생활 폐기물" in name:
        return WASTE_SCOPE_ALL_MUNICIPAL
    return WASTE_SCOPE_UNKNOWN


def covers_all_streams(scopes: Sequence[str]) -> bool:
    """True when the accepted contracts together cover 일반 + 음식물 + 재활용."""

    if WASTE_SCOPE_ALL_MUNICIPAL in scopes:
        return True
    return {
        WASTE_SCOPE_GENERAL_ONLY,
        WASTE_SCOPE_FOOD_ONLY,
        WASTE_SCOPE_RECYCLING_ONLY,
    }.issubset(set(scopes))


def classify_contract_accounting_basis(contract_name: str) -> str:
    """Accounting basis of one contract, from the tokens the source itself wrote.

    Two signals, both read from ``계약명`` and nothing else:

    ``COLLECTION_TRANSPORT_TOKENS``
        the indicator's own vocabulary — 수집, 운반, 수송, 청소.
    ``FACILITY_INBOUND_FEE_TOKENS`` / ``TREATMENT_ONLY_TOKENS``
        wording that names a different accounting basis.

    A contract is non-eligible only when it carries basis wording **and** no
    collection/transport wording at all. That asymmetry is the whole point: it
    removes 위탁처리 / 소각처리 / 선별처리 / 처리용역 / 처리대행 contracts while
    keeping 운반·처리 and 수집·운반 처리 대행용역, which really do pay for
    haulage. A bare ``처리`` never decides anything on its own — 253 of the 315
    audited 2024 contracts are ordinary 수집·운반 대행 contracts and 27 of them
    mention 처리 somewhere.

    A 반입수수료 is the one unconditional exclusion. It is a facility gate fee —
    the accounting basis of ``landfill_inbound_monthly``, which
    :mod:`waste_equity_backend.models.municipal_cost` states is never summed with
    this indicator — so no amount of surrounding collection wording can make it a
    collection-and-transport payment.

    Only ``계약명`` is read. ``처리방식`` and ``최종 처리시설`` describe where the
    waste *ends up*, not what was bought: in the audited delivery 45 of the 49
    rows whose 처리방식 is ``소각`` are ordinary 수집·운반 대행 contracts hauling
    to an incinerator, so treating either column as basis evidence would exclude
    real collection payment on a massive scale.
    """

    name = _text(contract_name)
    if any(token in name for token in FACILITY_INBOUND_FEE_TOKENS):
        return CONTRACT_BASIS_FACILITY_INBOUND_FEE
    if not any(token in name for token in TREATMENT_ONLY_TOKENS):
        return CONTRACT_BASIS_COLLECTION_TRANSPORT
    if any(token in name for token in COLLECTION_TRANSPORT_TOKENS):
        # Haulage bundled with the disposal of the same load. The transport
        # component is real collection-and-transport payment and is kept; the
        # bundling stays visible in the stored basis.
        return CONTRACT_BASIS_COLLECTION_TRANSPORT_WITH_TREATMENT
    return CONTRACT_BASIS_TREATMENT_SERVICE


def classify_payment_type(note: str) -> str:
    """Actual paid amount unless the source says otherwise, in its own words."""

    lowered = _text(note).lower()
    if any(marker in lowered for marker in CONTRACT_AWARD_MARKERS):
        return PAYMENT_CONTRACT_AWARD
    if any(_text(marker) in _text(note) for marker in BUDGET_ESTIMATE_MARKERS):
        return PAYMENT_BUDGET_ESTIMATE
    return PAYMENT_ACTUAL_PAID


def extract_contractor(contract_name: str, note: str) -> str | None:
    """Contractor from a source-labelled note field, or a company-suffixed tail."""

    note_text = _text(note)
    for label in _CONTRACTOR_LABELS:
        marker = f"{label}:"
        index = note_text.find(marker)
        if index < 0:
            index = note_text.find(label)
            if index < 0:
                continue
            start = index + len(label)
        else:
            start = index + len(marker)
        tail = note_text[start:].strip()
        candidate = re.split(r"[,;|]|\s{2,}|\.\s", tail, maxsplit=1)[0].strip(" .,;|")
        if candidate and _COMPANY_SUFFIX_PATTERN.search(candidate):
            return candidate[:200]

    name = _text(contract_name)
    if " - " in name:
        tail = name.rsplit(" - ", 1)[-1].strip()
        if tail and _COMPANY_SUFFIX_PATTERN.search(tail):
            return tail[:200]
    return None


def parse_period_label(
    label: str, reference_year: int
) -> tuple[datetime.date | None, datetime.date | None, bool]:
    """Parse a 년도 cell into ``(start, end, is_part_year)``.

    ``2024`` (or ``"2024"``) is a whole-year label and yields no dates — inventing
    2024-01-01/2024-12-31 would state a period the source never wrote. An explicit
    range such as ``2024.2.26.~4.5`` yields both dates and marks the contract
    part-year.
    """

    text = _text(label)
    if not text:
        return (None, None, False)
    if text == str(reference_year):
        return (None, None, False)
    if "~" not in text and "∼" not in text:
        # A single token that is neither the bare year nor a range: it declares
        # *something* narrower than the whole year, but no parseable endpoints.
        return (None, None, _PART_YEAR_PATTERN.match(text) is not None)

    left, _, right = text.replace("∼", "~").partition("~")
    start = _parse_date_token(left, reference_year)
    end = _parse_date_token(right, reference_year if start is None else start.year)
    return (start, end, True)


def _parse_date_token(token: str, default_year: int) -> datetime.date | None:
    match = _DATE_TOKEN_PATTERN.search(_text(token))
    if match is None:
        return None
    year = int(match.group(1)) if match.group(1) else default_year
    month, day = int(match.group(2)), int(match.group(3))
    try:
        return datetime.date(year, month, day)
    except ValueError:
        return None


def split_list_cell(raw: Any, separators: Sequence[str] = (";", "/")) -> tuple[str, ...]:
    """Split a semicolon/slash-delimited list cell, preserving each element verbatim."""

    text = _text(raw)
    if not text:
        return ()
    parts = [text]
    for separator in separators:
        expanded: list[str] = []
        for part in parts:
            expanded.extend(part.split(separator))
        parts = expanded
    return tuple(item.strip() for item in parts if item.strip())


# ---------------------------------------------------------------------------
# Workbook entry point
# ---------------------------------------------------------------------------


def parse_workbook(
    path: Path, *, dataset_role: str, reference_year: int, filename_annotation: str = ""
) -> ParsedWorkbook:
    """Parse one workbook read-only. The file is never modified or resaved."""

    try:
        workbook = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:  # pragma: no cover - unreadable file
        raise MunicipalCostParseError(f"could not open workbook: {exc}") from exc
    try:
        sheet = _primary_sheet(workbook)
        if sheet is None:
            return ParsedWorkbook(
                dataset_role=dataset_role,
                sheet_name=workbook.sheetnames[0] if workbook.sheetnames else "",
                used_range="",
                layout_family=LAYOUT_UNSUPPORTED,
                primary_classification=CLASS_EMPTY_OR_NO_DATA,
                reference_year=reference_year,
            )
        if dataset_role == DATASET_ROLE_B:
            return _parse_data_b(sheet, reference_year=reference_year)
        return _parse_data_a(
            sheet, reference_year=reference_year, filename_annotation=filename_annotation
        )
    finally:
        workbook.close()


def _primary_sheet(workbook: Workbook) -> Worksheet | None:
    """First sheet with any content; ``Sheet2``/``Sheet3`` are empty in every DATA_A."""

    for sheet in workbook.worksheets:
        if not (sheet.max_row and sheet.max_column):
            continue
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 5)):
            if any(cell.value is not None for cell in row):
                return sheet
    return None


# ---------------------------------------------------------------------------
# DATA_A
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class _DataAColumns:
    year: int
    organisation: int
    contract_name: int
    payment: int
    facility: int
    method: int
    note: int
    pair_columns: tuple[tuple[int, int], ...]
    ledger_month: int | None
    ledger_kind: int | None
    ledger_amount: int | None
    payment_header: str
    pair_headers: tuple[str, ...]


def _locate_data_a_columns(sheet: Worksheet, reference_year: int) -> _DataAColumns | None:
    header = {index: _text(cell.value) for index, cell in enumerate(sheet[1], start=1)}

    def find_exact(name: str) -> int | None:
        for index, text in header.items():
            if text == name:
                return index
        return None

    def find_any_exact(names: Sequence[str]) -> int | None:
        for name in names:
            found = find_exact(name)
            if found is not None:
                return found
        return None

    def find_prefix(prefix: str) -> int | None:
        for index, text in header.items():
            if text.startswith(prefix):
                return index
        return None

    year = find_any_exact(HEADER_YEAR_VARIANTS)
    organisation = find_exact(HEADER_ORGANISATION)
    contract_name = find_exact(HEADER_CONTRACT_NAME)
    payment = find_prefix(HEADER_PAYMENT_PREFIX)
    if payment is None:
        payment = find_any_exact(payment_header_variants(reference_year))
    facility = find_exact(HEADER_FACILITY)
    method = find_exact(HEADER_METHOD)
    note = find_exact(HEADER_NOTE)
    if None in (year, organisation, contract_name, payment, facility, method, note):
        return None
    assert payment is not None and facility is not None  # narrowed above
    if payment >= facility:
        return None

    # Everything strictly between the payment column and the trailing
    # 최종 처리시설 block is the [label, value] quantity pair region. Locating it
    # this way is what lets one code path handle the 9/11/13/14/15-column widths.
    span = list(range(payment + 1, facility))
    if len(span) % 2 != 0:
        return None
    pairs = tuple((span[i], span[i + 1]) for i in range(0, len(span), 2))

    return _DataAColumns(
        year=year,  # type: ignore[arg-type]
        organisation=organisation,  # type: ignore[arg-type]
        contract_name=contract_name,  # type: ignore[arg-type]
        payment=payment,
        facility=facility,
        method=method,  # type: ignore[arg-type]
        note=note,  # type: ignore[arg-type]
        pair_columns=pairs,
        ledger_month=find_exact(HEADER_LEDGER_MONTH),
        ledger_kind=find_exact(HEADER_LEDGER_KIND),
        ledger_amount=find_prefix(HEADER_LEDGER_AMOUNT_PREFIX),
        payment_header=header.get(payment, ""),
        pair_headers=tuple(header.get(column, "") for column, _ in pairs)
        + tuple(header.get(value, "") for _, value in pairs),
    )


def _classify_data_a_layout(sheet: Worksheet, columns: _DataAColumns) -> str:
    headers = " ".join(columns.pair_headers)
    if columns.ledger_month and columns.ledger_amount:
        return LAYOUT_DATA_A_CONTRACT_WITH_PAYMENT_LEDGER
    if "반출량(톤)" in headers and "반출량2(톤)" in headers:
        return LAYOUT_DATA_A_CONTRACT_TONNE_LABELLED_PAIRS
    if "월별 반출량 항목(" in headers:
        return LAYOUT_DATA_A_CONTRACT_DESTINATION_SPLIT
    if "의 수치)" in headers:
        return LAYOUT_DATA_A_CONTRACT_NUMBERED_PAIRS
    populated = [
        _text(sheet.cell(row, column).value)
        for row in range(2, sheet.max_row + 1)
        for pair in columns.pair_columns
        for column in pair
        if sheet.cell(row, column).value is not None
    ]
    if populated and all(_is_dash(text) for text in populated):
        return LAYOUT_DATA_A_CONTRACT_DASH_ONLY_QUANTITY
    return LAYOUT_DATA_A_CONTRACT_QUANTITY_PAIRS


def _parse_data_a(
    sheet: Worksheet, *, reference_year: int, filename_annotation: str
) -> ParsedWorkbook:
    columns = _locate_data_a_columns(sheet, reference_year)
    if columns is None:
        return ParsedWorkbook(
            dataset_role=DATASET_ROLE_A,
            sheet_name=sheet.title,
            used_range=str(sheet.dimensions),
            layout_family=LAYOUT_UNSUPPORTED,
            primary_classification=CLASS_UNSUPPORTED_LAYOUT,
            reference_year=reference_year,
            payment_limitation_reasons=[REASON_UNSUPPORTED_SOURCE_LAYOUT],
            warnings=["required DATA_A headers were not all found by name"],
        )

    parsed = ParsedWorkbook(
        dataset_role=DATASET_ROLE_A,
        sheet_name=sheet.title,
        used_range=str(sheet.dimensions),
        layout_family=_classify_data_a_layout(sheet, columns),
        primary_classification=CLASS_EMPTY_OR_NO_DATA,
        reference_year=reference_year,
    )

    annotation = _text(filename_annotation)
    filename_forces_general = any(
        marker in annotation for marker in FILENAME_GENERAL_ONLY_MARKERS
    ) and any(marker in annotation for marker in FILENAME_ONLY_MARKERS)
    if filename_forces_general:
        parsed.source_notes.append(
            "원본 파일명 주석이 생활폐기물(일반) 항목만 포함함을 명시함: " + annotation
        )

    blocks = list(_iter_contract_blocks(sheet, columns))
    organisations: list[str] = []
    contract_scopes: list[str] = []

    for block in blocks:
        row = block.start_row
        organisation = _text(sheet.cell(row, columns.organisation).value)
        if organisation and organisation not in organisations:
            organisations.append(organisation)

        note = _text(sheet.cell(row, columns.note).value)
        if note:
            parsed.source_notes.append(f"r{row}: {note}")

        contract_name = _text(sheet.cell(row, columns.contract_name).value)
        raw_payment = sheet.cell(row, columns.payment).value
        amount = _decimal(raw_payment)
        if amount is None:
            # A numeric cell delivered as formatted text ("6,556,861,120원").
            amount = parse_money_text(raw_payment)
        # The verbatim source text is kept whenever the cell was text, including
        # when it parsed cleanly, so the provenance of a text-formatted amount is
        # never lost.
        payment_text = _text(raw_payment) or None if isinstance(raw_payment, str) else None
        payment_type = classify_payment_type(note)
        # Three independent gates, in order: is there a number, is it a payment,
        # and is it a payment on *this* indicator's accounting basis.
        basis = classify_contract_accounting_basis(contract_name)
        basis_eligible = basis in NUMERATOR_ELIGIBLE_BASES
        eligible = amount is not None and payment_type == PAYMENT_ACTUAL_PAID and basis_eligible

        waste_scope = classify_contract_waste_scope(
            contract_name, filename_forces_general=filename_forces_general
        )
        # Waste-scope coverage describes what the *numerator* covers, so a
        # contract removed for its accounting basis does not get to supply a
        # stream it never paid to collect.
        if basis_eligible:
            contract_scopes.append(waste_scope)

        geographic_scope = (
            GEOGRAPHIC_SCOPE_SUB
            if SUB_MUNICIPAL_PATTERN.search(contract_name)
            else GEOGRAPHIC_SCOPE_WHOLE
        )

        period_label = _text(sheet.cell(row, columns.year).value) or None
        start, end, part_year = parse_period_label(period_label or "", reference_year)

        limitations: list[str] = []
        if part_year:
            limitations.append(REASON_PARTIAL_PERIOD_COVERAGE)
        if geographic_scope == GEOGRAPHIC_SCOPE_SUB:
            limitations.append(REASON_PARTIAL_GEOGRAPHIC_SCOPE)

        missing_months = _ledger_missing_months(sheet, columns, block, reference_year)
        if missing_months:
            limitations.append(REASON_PAYMENT_PERIOD_COVERAGE_INCOMPLETE)
            parsed.source_notes.append(
                f"r{row}: 지급 원장에 기록이 없는 월 — " + ", ".join(missing_months)
            )

        # Completeness describes the contract *record*, so it is decided before
        # the basis reason is attached: a 반입수수료 row is a complete record of a
        # real payment that simply belongs to another accounting basis.
        completeness = COMPLETENESS_COMPLETE
        if amount is None:
            completeness = COMPLETENESS_UNUSABLE
        elif limitations:
            completeness = COMPLETENESS_PARTIAL

        if not basis_eligible:
            limitations.append(REASON_NON_COLLECTION_TRANSPORT_BASIS)

        parsed.contracts.append(
            ParsedContract(
                source_row=row,
                source_row_end=block.end_row,
                contract_name=contract_name[:400],
                contractor_name=extract_contractor(contract_name, note),
                source_period_label=period_label,
                payment_amount_krw=amount,
                payment_source_text=payment_text[:200] if payment_text else None,
                payment_type=payment_type,
                accounting_basis=basis,
                is_primary_numerator_eligible=eligible,
                payment_period_start=start,
                payment_period_end=end,
                waste_scope=waste_scope,
                geographic_scope=geographic_scope,
                destination_names=split_list_cell(sheet.cell(row, columns.facility).value),
                treatment_methods=split_list_cell(sheet.cell(row, columns.method).value),
                completeness_status=completeness,
                limitation_reasons=tuple(dict.fromkeys(limitations)),
                source_note=note or None,
            )
        )

    parsed.source_municipality_names = tuple(organisations)

    # The source 합계 row is a reconciliation checksum only. It is never stored as
    # another contract and never added to the numerator alongside its own parts.
    total_row = _find_total_row(sheet, columns)
    if total_row is not None:
        raw_total = sheet.cell(total_row, columns.payment).value
        parsed.source_total_krw = _decimal(raw_total)
        if parsed.source_total_krw is None:
            # Same text-formatted-number case as the contract rows; parsing it is
            # what lets the reconciliation check actually run on those workbooks.
            parsed.source_total_krw = parse_money_text(raw_total)
        if parsed.source_total_krw is None:
            parsed.source_total_text = _text(raw_total) or None
    recomputed = [
        c.payment_amount_krw for c in parsed.contracts if c.payment_amount_krw is not None
    ]
    parsed.recomputed_total_krw = sum(recomputed, Decimal(0)) if recomputed else None
    if (
        parsed.source_total_krw is not None
        and parsed.recomputed_total_krw is not None
        and parsed.source_total_krw != parsed.recomputed_total_krw
    ):
        parsed.quantity_limitation_reasons.append(REASON_INCONSISTENT_TOTAL)
        parsed.warnings.append(
            f"source 합계 {parsed.source_total_krw} != recomputed {parsed.recomputed_total_krw}"
        )

    parsed.quantities = _parse_data_a_quantities(sheet, columns, blocks, parsed, reference_year)

    # A contract kept for provenance but excluded from the numerator is reported
    # at file level so the municipality's own reason codes say why its numerator
    # is smaller than the workbook's 합계 — and, when *every* contract is on
    # another basis, why there is no value at all.
    if any(
        contract.accounting_basis not in NUMERATOR_ELIGIBLE_BASES for contract in parsed.contracts
    ):
        parsed.payment_limitation_reasons.append(REASON_NON_COLLECTION_TRANSPORT_BASIS)

    # Waste scope is judged for the municipality as a whole: 미추홀구's separate
    # 일반 / 재활용 / 음식물 contracts together cover every stream, so the file is
    # not partial even though no single contract is. Judged over the accepted
    # contracts only — with none of them left there is no scope to be partial
    # about, and MISSING_PAYMENT is the honest finding rather than a claim that
    # some streams were covered.
    if contract_scopes and not covers_all_streams(contract_scopes):
        parsed.payment_limitation_reasons.append(REASON_PARTIAL_WASTE_SCOPE)

    parsed.primary_classification = _classify_data_a_content(parsed)
    return parsed


@dataclass(frozen=True)
class _ContractBlock:
    start_row: int
    end_row: int


def _find_total_row(sheet: Worksheet, columns: _DataAColumns) -> int | None:
    for row in range(2, sheet.max_row + 1):
        if _text(sheet.cell(row, columns.contract_name).value) in TOTAL_ROW_TOKENS:
            return row
    return None


def _iter_contract_blocks(sheet: Worksheet, columns: _DataAColumns) -> Iterator[_ContractBlock]:
    """Yield each contract header row with the extent of the block beneath it."""

    header_rows: list[int] = []
    boundary = sheet.max_row + 1
    for row in range(2, sheet.max_row + 1):
        name = _text(sheet.cell(row, columns.contract_name).value)
        if not name:
            continue
        if name == HEADER_CONTRACT_NAME:
            # A repeated header row inside the data region: 미추홀구 restates the
            # whole header on row 2 with a different payment-column spelling. It is
            # a header, not a contract literally named "계약명", and reading it as
            # one would invent a contract and pollute the 기관명 signal used to
            # resolve the municipality.
            continue
        if name in TOTAL_ROW_TOKENS:
            boundary = min(boundary, row)
            continue
        header_rows.append(row)
    for index, row in enumerate(header_rows):
        following = header_rows[index + 1] if index + 1 < len(header_rows) else boundary
        yield _ContractBlock(start_row=row, end_row=max(row, min(following, boundary) - 1))


def _ledger_missing_months(
    sheet: Worksheet,
    columns: _DataAColumns,
    block: _ContractBlock,
    reference_year: int,
) -> list[str]:
    """Months of the reference year with no entry in a per-contract payment ledger.

    Derived from the ledger itself rather than from prose, so the finding is
    reproducible: 계양구's four contracts each record only 2·3·9·10·11·12월.
    """

    if columns.ledger_month is None:
        return []
    present: set[str] = set()
    for row in range(block.start_row, block.end_row + 1):
        token = _text(sheet.cell(row, columns.ledger_month).value)
        match = _MONTH_ISO_PATTERN.match(token)
        if match and int(match.group(1)) == reference_year:
            present.add(f"{reference_year:04d}-{int(match.group(2)):02d}")
    if not present:
        return []
    expected = {f"{reference_year:04d}-{month:02d}" for month in range(1, 13)}
    return sorted(expected - present)


def _block_signature(
    sheet: Worksheet, columns: _DataAColumns, block: _ContractBlock
) -> tuple[tuple[str, str], ...]:
    """(label, raw-value) signature of a contract's quantity block, in sheet order."""

    signature: list[tuple[str, str]] = []
    for row in range(block.start_row, block.end_row + 1):
        for label_column, value_column in columns.pair_columns:
            label_raw = sheet.cell(row, label_column).value
            value_raw = sheet.cell(row, value_column).value
            if label_raw is None and value_raw is None:
                continue
            signature.append((_text(label_raw), repr(value_raw)))
    return tuple(signature)


def _parse_data_a_quantities(
    sheet: Worksheet,
    columns: _DataAColumns,
    blocks: Sequence[_ContractBlock],
    parsed: ParsedWorkbook,
    reference_year: int,
) -> list[ParsedQuantity]:
    signatures = [_block_signature(sheet, columns, block) for block in blocks]

    # A city-wide block copied verbatim under several contracts is **one** logical
    # municipal series, not N per-contract measurements. Grouping by signature —
    # rather than requiring every block in the file to match — is what catches
    # 광명시, where seven 구역 contracts share the city-wide series and an eighth
    # 민간소각 contract carries its own. Requiring a numeric value in the shared
    # signature keeps a repeated "미제공" placeholder (군포시) from being promoted
    # to a municipal total.
    groups: dict[tuple[tuple[str, str], ...], list[int]] = {}
    for index, signature in enumerate(signatures):
        groups.setdefault(signature, []).append(index)
    repeated_groups = {
        signature: members
        for signature, members in groups.items()
        if len(members) > 1
        and signature
        and any(_decimal_from_repr(value) is not None for _, value in signature)
    }
    # Index of the block that represents each repeated group (the first
    # occurrence), and the group's full membership for provenance.
    representative_of: dict[int, list[int]] = {}
    suppressed: set[int] = set()
    for members in repeated_groups.values():
        representative_of[members[0]] = members
        suppressed.update(members[1:])

    quantities: list[ParsedQuantity] = []

    for block_index, block in enumerate(blocks):
        if block_index in suppressed:
            continue
        group_members = representative_of.get(block_index)
        repeated = group_members is not None
        repetition_offsets: list[int] = (
            [blocks[member].start_row - block.start_row for member in group_members]
            if group_members is not None
            else [0]
        )
        contract = parsed.contracts[block_index] if block_index < len(parsed.contracts) else None
        scope = contract.waste_scope if contract else WASTE_SCOPE_UNKNOWN
        for row in range(block.start_row, block.end_row + 1):
            for label_column, value_column in columns.pair_columns:
                label_raw = sheet.cell(row, label_column).value
                value_raw = sheet.cell(row, value_column).value
                if label_raw is None and value_raw is None:
                    continue
                label = _text(label_raw)
                value, state = classify_value(value_raw, label=label)
                month, period = parse_label_period(label, reference_year)
                limitations: list[str] = []
                if month is not None and not month.startswith(f"{reference_year:04d}"):
                    limitations.append(REASON_MIXED_REFERENCE_YEARS)
                destination = label.rsplit(":", 1)[0].strip() if ":" in label else (label or None)
                if repeated:
                    rows = tuple(row + offset for offset in repetition_offsets)
                    attribution = ATTRIBUTION_MUNICIPAL_TOTAL_REPEATED
                    contract_index = None
                else:
                    rows = (row,)
                    attribution = ATTRIBUTION_PER_CONTRACT
                    contract_index = block_index
                quantities.append(
                    ParsedQuantity(
                        source_row=row,
                        source_column=value_column,
                        source_label=label[:200],
                        source_repetition_rows=rows,
                        contract_index=contract_index,
                        reference_month=month,
                        quantity_period=period,
                        waste_category=classify_waste_category(label, scope),
                        destination_name=(destination[:300] if destination else None),
                        treatment_method=None,
                        quantity_value=value,
                        value_state=state,
                        attribution=attribution,
                        limitation_reasons=tuple(limitations),
                    )
                )

    for members in repeated_groups.values():
        member_rows = ", ".join(f"r{blocks[member].start_row}" for member in members)
        parsed.source_notes.append(
            f"동일한 월별 반출량 블록이 {len(members)}개 계약 행({member_rows}) 아래에 반복되어 "
            "있어 1회만 지자체 단위 합계로 저장하고 반복 위치는 모두 기록했습니다."
        )
    for quantity in quantities:
        for reason in quantity.limitation_reasons:
            if reason not in parsed.quantity_limitation_reasons:
                parsed.quantity_limitation_reasons.append(reason)

    # The source itself converting '-' to 0 makes those zeros unverifiable; it is
    # recorded on the quantity side only and never blocks the payment indicator.
    joined_notes = " ".join(parsed.source_notes)
    if "'-' 표기는 0으로 처리" in joined_notes or "- 표기는 0으로 처리" in joined_notes:
        if REASON_AMBIGUOUS_SOURCE_MEANING not in parsed.quantity_limitation_reasons:
            parsed.quantity_limitation_reasons.append(REASON_AMBIGUOUS_SOURCE_MEANING)
    return quantities


def _decimal_from_repr(value_repr: str) -> Decimal | None:
    try:
        return Decimal(value_repr)
    except (InvalidOperation, ValueError):
        return None


def _classify_data_a_content(parsed: ParsedWorkbook) -> str:
    # What the workbook *delivered*, not what survived numerator eligibility: a
    # file of 반입수수료 rows carries 18.3 billion KRW of real contracts and must
    # never be filed as EMPTY_OR_NO_DATA next to the contract rows it produced.
    # Eligibility is answered per contract by is_primary_numerator_eligible and
    # per municipality by the indicator's reason codes.
    has_payment = any(contract.payment_amount_krw is not None for contract in parsed.contracts)
    has_quantity = parsed.has_numeric_quantity
    if not parsed.contracts and not has_quantity:
        return CLASS_EMPTY_OR_NO_DATA
    if has_payment and REASON_PARTIAL_WASTE_SCOPE in parsed.payment_limitation_reasons:
        return CLASS_DATA_A_PARTIAL_WASTE_SCOPE
    if has_payment and has_quantity:
        return CLASS_DATA_A_PAYMENT_AND_QUANTITY
    if has_payment:
        return CLASS_DATA_A_PAYMENT_ONLY
    if has_quantity:
        return CLASS_DATA_A_QUANTITY_ONLY
    return CLASS_EMPTY_OR_NO_DATA


# ---------------------------------------------------------------------------
# DATA_B
# ---------------------------------------------------------------------------


def _parse_data_b(sheet: Worksheet, *, reference_year: int) -> ParsedWorkbook:
    header = {index: _text(cell.value) for index, cell in enumerate(sheet[1], start=1)}
    category_column = next(
        (index for index, text in header.items() if text == DATA_B_HEADER_CATEGORY), None
    )
    month_columns: dict[int, int] = {}
    total_column: int | None = None
    for index, text in header.items():
        match = _MONTH_KO_PATTERN.match(text)
        if match and text.endswith("월"):
            month_columns[int(match.group(1))] = index
        elif text == DATA_B_HEADER_TOTAL:
            total_column = index

    parsed = ParsedWorkbook(
        dataset_role=DATASET_ROLE_B,
        sheet_name=sheet.title,
        used_range=str(sheet.dimensions),
        layout_family=LAYOUT_DATA_B_MONTHLY_CATEGORY_GRID,
        primary_classification=CLASS_DATA_B_QUANTITY_VALIDATION,
        reference_year=reference_year,
    )
    if category_column is None or len(month_columns) != 12 or total_column is None:
        parsed.layout_family = LAYOUT_UNSUPPORTED
        parsed.primary_classification = CLASS_UNSUPPORTED_LAYOUT
        parsed.payment_limitation_reasons.append(REASON_UNSUPPORTED_SOURCE_LAYOUT)
        parsed.warnings.append("DATA_B grid headers 구분 / 1월…12월 / 계 were not all found")
        return parsed

    for row in range(2, sheet.max_row + 1):
        label = _text(sheet.cell(row, category_column).value)
        if not label:
            continue
        if label.startswith("※"):
            parsed.source_notes.append(f"r{row}: {label}")
            continue
        if label not in DATA_B_CATEGORY_ROWS:
            parsed.source_notes.append(f"r{row}: {label}")
            continue

        category = DATA_B_CATEGORY_TO_WASTE[label]
        monthly_numeric: list[Decimal] = []
        monthly_states: list[str] = []
        for month in range(1, 13):
            column = month_columns[month]
            value, state = classify_value(sheet.cell(row, column).value)
            monthly_states.append(state)
            if value is not None:
                monthly_numeric.append(value)
            parsed.quantities.append(
                ParsedQuantity(
                    source_row=row,
                    source_column=column,
                    source_label=f"{label}:{month}월",
                    source_repetition_rows=(row,),
                    contract_index=None,
                    reference_month=f"{reference_year:04d}-{month:02d}",
                    quantity_period=QUANTITY_PERIOD_MONTHLY,
                    waste_category=category,
                    destination_name=None,
                    treatment_method=None,
                    quantity_value=value,
                    value_state=state,
                    attribution=ATTRIBUTION_MUNICIPAL_TOTAL_SINGLE,
                )
            )

        raw_total = sheet.cell(row, total_column).value
        total_value, total_state = classify_value(raw_total)
        limitations: list[str] = []
        # A cached =SUM() over twelve '-' placeholders caches 0. That zero is an
        # artefact, not a measured zero: it is stored as SOURCE_DASH_NO_DATA with
        # a NULL value, exactly like the dashes it summed.
        if (
            total_value is not None
            and total_value == 0
            and not monthly_numeric
            and all(state == VALUE_SOURCE_DASH_NO_DATA for state in monthly_states)
        ):
            total_value, total_state = None, VALUE_SOURCE_DASH_NO_DATA
            limitations.append(REASON_ZERO_TOTAL_QUANTITY)
        elif total_value is not None and monthly_numeric:
            # Both sides are already at the source's declared precision, so this
            # is an exact comparison, not a tolerance.
            recomputed = quantize_quantity(sum(monthly_numeric, Decimal(0)))
            if recomputed != total_value:
                limitations.append(REASON_INCONSISTENT_TOTAL)
                parsed.warnings.append(
                    f"r{row} {label}: cached 계 {total_value} != recomputed {recomputed}"
                )
        parsed.quantities.append(
            ParsedQuantity(
                source_row=row,
                source_column=total_column,
                source_label=f"{label}:계",
                source_repetition_rows=(row,),
                contract_index=None,
                reference_month=None,
                quantity_period=QUANTITY_PERIOD_ANNUAL,
                waste_category=category,
                destination_name=None,
                treatment_method=None,
                quantity_value=total_value,
                value_state=total_state,
                attribution=ATTRIBUTION_MUNICIPAL_TOTAL_SINGLE,
                limitation_reasons=tuple(limitations),
            )
        )
        for reason in limitations:
            if reason not in parsed.quantity_limitation_reasons:
                parsed.quantity_limitation_reasons.append(reason)

    if not parsed.has_numeric_quantity:
        parsed.primary_classification = CLASS_EMPTY_OR_NO_DATA
        parsed.quantity_limitation_reasons.append(REASON_MISSING_QUANTITY)
    return parsed
